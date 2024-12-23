

from aiogram import Router, Bot, F
from aiogram.types import Message
from aiogram.fsm.context import FSMContext
from aiogram.filters import CommandStart, StateFilter
from aiogram.fsm.state import State, StatesGroup

from database import requests as rq
from config_data.config import Config, load_config
from utils.error_handling import error_handler

from openpyxl import load_workbook
import logging
import requests

config: Config = load_config()
router = Router()
router.message.filter(F.chat.type == "private")


class User(StatesGroup):
    article = State()


@router.message(CommandStart())
@error_handler
async def start(message: Message, state: FSMContext, bot: Bot) -> None:
    """
    Запуск бота
    :param message:
    :param state:
    :param bot:
    :return:
    """
    logging.info('start')
    tg_id = message.chat.id
    await state.set_state(state=None)
    user = await rq.get_user(tg_id=tg_id)
    if not user:
        if message.from_user.username:
            username = message.from_user.username
        else:
            username = "useraname"
        data = {"tg_id": tg_id, "username": username}
        await rq.add_user(tg_id=tg_id, data=data)
    await message.answer(text='Добро пожаловать! Вас приветствует виртуальный помощник Ирина.\n'
                              'Для получения остатка товаров отправьте его артикул из каталога товара'
                              ' (например, 4991210)')
    # await message.answer_photo(photo='AgACAgIAAxkBAAMCZzzHbJTLg_V6cHFsgot9O1CszfQAAqLtMRvVT-hJ2sU--LfaV_gBAAMCAAN4AAM2BA',
    #                            caption='Добро пожаловать! Вас приветствует виртуальный помощник Ирина.\n'
    #                                    'Для получения остатка товаров отправьте его артикул из каталога товара'
    #                                    ' (например, 4991210)')
    await state.set_state(User.article)


@router.message(StateFilter(User.article))
async def get_article(message: Message, state: FSMContext, bot: Bot):
    logging.info('get_article')
    # await message.answer(text='Я уже побежала на склад проверять наличие товара по вашему запросу, минутку... ⏳')
    if message.text:
        text = '<b>Количество товара на складах:</b>\n\n'
        flag = 0
        product_spb = await rq.get_product_spb(article=message.text.lower())
        if product_spb:
            leftovers_spb = product_spb.count
            if leftovers_spb == 10:
                text += f'<b>СПб:</b>\n' \
                        f'артикул - <i>{product_spb.article}</i>\n' \
                        f'наименование - <i><u>{product_spb.title}</u></i>\n' \
                        f'количество - более 10 шт, РРЦ {product_spb.amount_rrc} ₽.\n\n'
            else:
                text += f'<b>СПб:</b>\n' \
                        f'артикул - <i>{product_spb.article}</i>\n' \
                        f'наименование - <i><u>{product_spb.title}</u></i>\n' \
                        f'количество - {leftovers_spb} шт, РРЦ {product_spb.amount_rrc} ₽.\n\n'
            flag = 1
        product_msk = await rq.get_product_msk(article=message.text.lower())
        if product_msk:
            leftovers_msk = product_msk.count
            if leftovers_msk == 10:
                text += f'<b>Мск:</b>\n' \
                        f'артикул - <i>{product_msk.article}</i>\n' \
                        f'наименование - <i><u>{product_msk.title}</u></i>\n' \
                        f'количество - более 10 шт, РРЦ {product_msk.amount_rrc} ₽.\n\n'
            else:
                text += f'<b>Мск:</b>\n' \
                        f'артикул - <i>{product_msk.article}</i>\n' \
                        f'наименование - <i><u>{product_msk.title}</u></i>\n' \
                        f'количество - {leftovers_msk} шт, РРЦ {product_msk.amount_rrc} ₽.\n\n'
            flag = 1

        if flag:
            await message.answer(text=text)
        else:
            await message.answer(text=f'Артикул не найден')
    else:
        await message.answer(text='Некорректный номер артикула, повторите ввод')


async def updating_data():
    logging.info('updating_data')

    await rq.delete_spb()
    await rq.delete_msk()

    dls = "https://omoikiri.ru/ostatki/Ostatki_Omoikiri_SPb%20(XLSX).xlsx"
    resp = requests.get(dls)
    output = open('SPb.xlsx', 'wb')
    output.write(resp.content)
    output.close()

    dls = "https://omoikiri.ru/ostatki/Ostatki_Omoikiri_Msk%20(XLSX).xlsx"
    resp = requests.get(dls)
    output = open('Msk.xlsx', 'wb')
    output.write(resp.content)
    output.close()

    wb_spb = load_workbook('SPb.xlsx')
    sheet_spb = wb_spb.active
    search_col = 0
    col_title = 1
    col_amount_rrc = 1
    col_article = 1
    col_count = 1
    for row_num in range(1, sheet_spb.max_row):
        if sheet_spb.cell(row=row_num, column=1).value == 'Склад':
            col_article = 1
            for col, cell in enumerate(sheet_spb[row_num]):
                if cell.value == 'Номенклатура':
                    col_title = col + 1
                elif cell.value == 'РРЦ':
                    col_amount_rrc = col + 1
                elif cell.value == 'Сейчас':
                    col_count = col + 1
            search_col = 1
        elif search_col == 0:
            continue
        if search_col:
            data = {'article': sheet_spb.cell(row=row_num, column=col_article).value.lower() if sheet_spb.cell(row=row_num, column=col_article).value else "0",
                    'title': sheet_spb.cell(row=row_num, column=col_title).value if sheet_spb.cell(row=row_num, column=col_title).value else "0",
                    'amount_rrc': sheet_spb.cell(row=row_num, column=col_amount_rrc).value if sheet_spb.cell(row=row_num, column=col_amount_rrc).value else 0,
                    'count': sheet_spb.cell(row=row_num, column=col_count).value if sheet_spb.cell(row=row_num, column=col_count).value else 0}
            await rq.add_product_spb(data=data)

    wb_msk = load_workbook('Msk.xlsx')
    sheet_msk = wb_msk.active
    search_col = 0
    col_title = 1
    col_amount_rrc = 1
    col_count = 1
    for row_num in range(1, sheet_msk.max_row):
        if sheet_msk.cell(row=row_num, column=1).value == 'Склад':
            col_article = 1
            for col, cell in enumerate(sheet_msk[row_num]):
                if cell.value == 'Номенклатура':
                    col_title = col + 1
                elif cell.value == 'РРЦ':
                    col_amount_rrc = col + 1
                elif cell.value == 'Сейчас':
                    col_count = col + 1
            search_col = 1
        elif search_col == 0:
            continue
        if search_col:
            data = {'article': sheet_msk.cell(row=row_num, column=col_article).value.lower() if sheet_msk.cell(row=row_num, column=col_article).value else "0",
                    'title': sheet_msk.cell(row=row_num, column=col_title).value if sheet_msk.cell(row=row_num, column=col_title).value else "0",
                    'amount_rrc': sheet_msk.cell(row=row_num, column=col_amount_rrc).value if sheet_msk.cell(row=row_num, column=col_amount_rrc).value else 0,
                    'count': sheet_msk.cell(row=row_num, column=col_count).value if sheet_msk.cell(row=row_num, column=col_count).value else 0}
            await rq.add_product_msk(data=data)
